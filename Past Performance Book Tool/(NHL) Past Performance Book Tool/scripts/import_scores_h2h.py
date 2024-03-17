import os
import csv
import pandas as pd
import openpyxl as px

# Function to calculate points scored by each team
def calculate_points(data, team1, team2):
    team1_points = 0
    team2_points = 0
    games_calculated = 0

    for row in data:
        if (row["Visitor"] == team1 or row["Home"] == team1) and (row["Visitor"] == team2 or row["Home"] == team2):
            team1_points += float(row["G"])
            team2_points += float(row["G.1"])
            games_calculated += 1

    return team1_points, team2_points, games_calculated

# Function to update 'container.xlsx' with averages
def update_excel_with_averages(CSV_PATH, EXCEL_PATH):
    # Read data from the CSV file
    with open(CSV_PATH, newline='') as file:
        reader = csv.DictReader(file)
        data = list(reader)

    # Read team names from the Excel file
    teams_df = pd.read_excel(EXCEL_PATH)
    team_names = teams_df['Match-up'].tolist()

    # Calculate averages
    averages = {}
    for i in range(0, len(team_names), 2):
        team1 = team_names[i]
        team2 = team_names[i + 1]
        
        team1_points, team2_points, games_calculated = calculate_points(data, team1, team2)
        
        if games_calculated > 0:
            average_team1_points = team1_points / games_calculated
            average_team2_points = team2_points / games_calculated
        else:
            average_team1_points = ""
            average_team2_points = ""
        
        averages[team1] = average_team1_points
        averages[team2] = average_team2_points

    # Update 'container.xlsx'
    book = px.load_workbook(EXCEL_PATH)
    sheet = book.active

    # Read and store the header row
    header = [cell.value for cell in sheet[1]]

    # Clear the specified column (24th column)
    for cell in sheet.iter_cols(min_col=24, max_col=24):
        cell[0].value = None

    # Write the new data to the specified column
    for i, team_name in enumerate(team_names, start=2):
        avg = averages.get(team_name, "")
        sheet.cell(row=i, column=24, value=avg)

    # Reinsert the header row
    for i, header_value in enumerate(header, start=1):
        sheet.cell(row=1, column=i, value=header_value)

    book.save(EXCEL_PATH)

def main():
    CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'pp.csv')
    EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'container.xlsx')

    # Call the function to update averages in 'container.xlsx'
    update_excel_with_averages(CSV_PATH, EXCEL_PATH)

if __name__ == "__main__":
    main()
    print("Completed")