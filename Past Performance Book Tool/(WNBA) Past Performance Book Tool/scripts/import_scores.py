import pandas as pd
import os
import openpyxl as px

# Sum-up the total of each teams past performance
def calculate_averages(team_data_pts, team_data_allowed):
    recent_3_pts = team_data_pts.head(3)
    recent_10_pts = team_data_pts.head(10)
    recent_36_pts = team_data_pts.head(36)

    recent_3_allowed = team_data_allowed.head(3)
    recent_10_allowed = team_data_allowed.head(10)
    recent_36_allowed = team_data_allowed.head(36)

    average_3_pts = recent_3_pts.mean()
    average_10_pts = recent_10_pts.mean()
    average_36_pts = recent_36_pts.mean()

    average_3_allowed = recent_3_allowed.mean()
    average_10_allowed = recent_10_allowed.mean()
    average_36_allowed = recent_36_allowed.mean()

    return (
        f"{average_3_pts:.2f}" if not pd.isna(average_3_pts) else "",
        f"{average_10_pts:.2f}" if not pd.isna(average_10_pts) else "",
        f"{average_36_pts:.2f}" if not pd.isna(average_36_pts) else "",
        f"{average_3_allowed:.2f}" if not pd.isna(average_3_allowed) else "",
        f"{average_10_allowed:.2f}" if not pd.isna(average_10_allowed) else "",
        f"{average_36_allowed:.2f}" if not pd.isna(average_36_allowed) else ""
    )

# Write data to the container
def main():
    CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'pp.csv')
    EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'container.xlsx')

    df = pd.read_csv(CSV_PATH)
    teams_df = pd.read_excel(EXCEL_PATH)
    team_names = teams_df['Match-up'].tolist()

    recent_averages = {}

    for team in team_names:
        wins_pts = df[df['Visitor/Neutral'] == team]['PTS']
        wins_allowed = df[df['Visitor/Neutral'] == team]['PTS.1']
        losses_pts = df[df['Home/Neutral'] == team]['PTS.1']
        losses_allowed = df[df['Home/Neutral'] == team]['PTS']

        team_data_pts = pd.concat([wins_pts, losses_pts])
        team_data_allowed = pd.concat([wins_allowed, losses_allowed])

        team_data_pts = team_data_pts.sort_index(ascending=False)
        team_data_allowed = team_data_allowed.sort_index(ascending=False)

        recent_averages[team] = calculate_averages(team_data_pts, team_data_allowed)

    book = px.load_workbook(EXCEL_PATH)
    sheet = book.active

    # Read and store the header row
    header = [cell.value for cell in sheet[1]]

    # Clear the specified columns
    for col in range(25, 31):
        for cell in sheet.iter_cols(min_col=col, max_col=col):
            cell[0].value = None

    # Write the new data to the specified columns, replacing NaN with an empty string
    for i, team_name in enumerate(team_names, start=2):
        averages = recent_averages[team_name]
        for col, avg in zip(range(25, 31), averages):
            sheet.cell(row=i, column=col, value=avg)

    # Reinsert the header row
    for i, header_value in enumerate(header, start=1):
        sheet.cell(row=1, column=i, value=header_value)

    book.save(EXCEL_PATH)

if __name__ == '__main__':
    main()
    print("Completed")
